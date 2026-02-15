#!/usr/bin/env python3
"""
web_scan_enhanced.py

Enhanced lightweight web server scanner:
- Identifies server software (Server / X-Powered-By) and attempts OS detection
- Uses python-nmap if available (falls back to shelling out to nmap)
- Supports nmap "profiles" including stealth/paranoid/decoy to reduce detection (heuristics)
- Performs heuristic CVE lookups using CIRCL cve.circl.lu API for detected products/versions
- Checks common sensitive paths, security headers, HTTP methods, directory listings, TLS info
- Outputs report as HTML, XLSX, or JSON

Usage:
    python web_scan_enhanced.py https://example.com --output report.html
    python web_scan_enhanced.py https://example.com --output report.json --nmap-profile stealth
    python web_scan_enhanced.py https://example.com --output report.xlsx --no-nmap

Legal: Only run this against systems you own or are explicitly authorized to test.

Author: Danny Vargas
"""

from __future__ import annotations
import argparse
import requests
import urllib.parse
import re
import socket
import ssl
import datetime
import json
import os
import subprocess
import shutil
import sys
from typing import Dict, Any, List, Optional, Tuple

# Optional libs
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

# try python-nmap
try:
    import nmap  # python-nmap
except Exception:
    nmap = None

# Constants
COMMON_SENSITIVE_PATHS = [
    '/', 
    '/.git/', 
    '/.env', 
    '/.env.example', 
    '/robots.txt', 
    '/backup.zip',
    '/backup.tar.gz', 
    '/.htaccess', 
    '/phpinfo.php', 
    '/info.php', 
    '/admin/',
    '/config.php', 
    '/wp-login.php', 
    '/server-status', 
    '/.DS_Store'
]

SECURITY_HEADERS = [
    'Strict-Transport-Security',
    'Content-Security-Policy',
    'X-Frame-Options',
    'X-Content-Type-Options',
    'Referrer-Policy',
    'Permissions-Policy',
    'Expect-CT'
]

CIRCL_CVE_SEARCH = "https://cve.circl.lu/api/search/{}"  # search by keyword

# ---------------------------
# Helpers
# ---------------------------
def merge_ports(*port_sets):
    """
    Accepts strings, lists, or None.
    Produces ONE valid comma-separated port string for nmap.
    """
    ports = []

    for p in port_sets:
        if not p:
            continue
        if isinstance(p, str):
            parts = p.split(",")
            ports.extend([x.strip() for x in parts if x.strip()])
        elif isinstance(p, list):
            ports.extend([str(x).strip() for x in p if str(x).strip()])

    # make unique & numeric sorted
    ports = sorted(set(ports), key=lambda x: int(x))
    return ",".join(ports)


def norm_url(url: str) -> str:
    if not re.match(r'^https?://', url):
        url = 'http://' + url
    return url.rstrip('/')

def safe_request(url: str, timeout: int = 12, allow_redirects: bool = True):
    headers = {'User-Agent': 'web-scan-enhanced/1.0 (+https://example.com)'}
    try:
        r = requests.get(url, timeout=timeout, headers=headers, allow_redirects=allow_redirects, verify=False)
        return r
    except Exception as e:
        return e

def parse_host_port(parsed: urllib.parse.ParseResult) -> Tuple[str,int]:
    port = parsed.port
    if not port:
        port = 443 if parsed.scheme == 'https' else 80
    host = parsed.hostname
    return host, port

def tls_info(host: str, port: int, timeout: int = 6) -> Dict[str,Any]:
    info = {}
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with socket.create_connection((host, port), timeout=timeout) as sock:
            with ctx.wrap_socket(sock, server_hostname=host) as ssock:
                cert = ssock.getpeercert()
                info['protocol'] = ssock.version()
                info['cipher'] = ssock.cipher()
                info['cert_subject'] = cert.get('subject', ())
                info['cert_issuer'] = cert.get('issuer', ())
                info['notBefore'] = cert.get('notBefore')
                info['notAfter'] = cert.get('notAfter')
    except Exception as e:
        info['error'] = str(e)
    return info

# ---------------------------
# HTTP checks
# ---------------------------
def check_headers_and_server(url: str) -> Dict[str,Any]:
    r = safe_request(url)
    result = {}
    if isinstance(r, Exception):
        result['error'] = str(r)
        return result
    result['status_code'] = r.status_code
    result['final_url'] = r.url
    hdrs = dict(r.headers)
    result['headers'] = hdrs
    result['server_header'] = hdrs.get('Server')
    result['x_powered_by'] = hdrs.get('X-Powered-By') or hdrs.get('x-powered-by')
    return result

def check_security_headers(headers: Dict[str,str]) -> Dict[str,Any]:
    present = {}
    missing = []
    for h in SECURITY_HEADERS:
        found = None
        for k in headers.keys():
            if k.lower() == h.lower():
                found = headers[k]
                break
        if found is not None:
            present[h] = found
        else:
            missing.append(h)
    return {'present': present, 'missing': missing}

def check_http_methods(url: str) -> Dict[str,Any]:
    headers = {'User-Agent': 'web-scan-enhanced/1.0'}
    try:
        r = requests.options(url, timeout=10, headers=headers, verify=False)
        allow = r.headers.get('Allow') or r.headers.get('allow')
        methods = []
        if allow:
            methods = [m.strip() for m in allow.split(',')]
        else:
            for m in ['GET', 'HEAD', 'POST', 'TRACE']:
                try:
                    rr = requests.request(m, url, timeout=8, headers=headers, verify=False, allow_redirects=False)
                    methods.append((m, rr.status_code))
                except Exception:
                    methods.append((m, 'error'))
        return {'allow_header': allow, 'methods_detected': methods}
    except Exception as e:
        return {'error': str(e)}

def check_common_paths(base_url: str, paths: List[str], fast: bool=False) -> Dict[str,Any]:
    results = []
    headers = {'User-Agent': 'web-scan-enhanced/1.0'}
    for p in paths:
        full = urllib.parse.urljoin(base_url + '/', p.lstrip('/'))
        try:
            r = requests.get(full, timeout=8, headers=headers, verify=False, allow_redirects=True)
            status = r.status_code
            snippet = (r.text[:400] + '...') if isinstance(r.text, str) else ''
            interesting = status in (200,403,401,500) or ('Index of /' in r.text if isinstance(r.text,str) else False) or ('.git' in p and status==200)
            results.append({'path': p, 'url': full, 'status': status, 'interesting': interesting, 'snippet': snippet})
        except Exception as e:
            results.append({'path': p, 'url': full, 'status': 'error', 'error': str(e)})
        if fast:
            # continue but don't delay — fast mode just keeps request list small (caller can reduce list)
            continue
    return {'results': results}

def check_directory_listing(base_url: str) -> Dict[str,Any]:
    dirs = ['', '/', '/images/', '/uploads/', '/static/', '/wp-content/']
    findings = []
    headers = {'User-Agent': 'web-scan-enhanced/1.0'}
    for d in dirs:
        full = urllib.parse.urljoin(base_url + '/', d.lstrip('/'))
        try:
            r = requests.get(full, timeout=8, headers=headers, verify=False, allow_redirects=True)
            if r.status_code == 200 and ('Index of /' in r.text or '<title>Index of' in r.text or 'Parent Directory' in r.text):
                findings.append({'dir': d or '/', 'url': full, 'listing': True})
            else:
                findings.append({'dir': d or '/', 'url': full, 'listing': False, 'status': r.status_code})
        except Exception as e:
            findings.append({'dir': d or '/', 'url': full, 'error': str(e)})
    return {'results': findings}

# ---------------------------
# Nmap integration and profiles
# ---------------------------
def build_nmap_args(host: str, ports: Optional[str], profile: str) -> List[str]:
    """
    Build a list of nmap arguments based on profile.
    Profiles:
      - normal: -sV -p <ports> (default)
      - stealth: -sS -sV -Pn -T1 --scan-delay 200ms --max-retries 2 --data-length 24
      - paranoid: -sS -sV -Pn -T0 --scan-delay 1s --max-retries 5 --data-length 32
      - decoy: -sS -sV -Pn -T2 --data-length 24 --decoy 192.0.2.1,198.51.100.2 (example IPs)
      - custom: you can pass additional nmap_args via function (not implemented in CLI)
    Note: many options require privileges (raw socket scans).
    """
    base = []
    if profile == 'normal':
        base = ['-sV']
    elif profile == 'stealth':
        base = ['-sS', '-sV', '-Pn', '-T1', '--scan-delay', '200ms', '--max-retries', '2', '--data-length', '24']
    elif profile == 'paranoid':
        base = ['-sS', '-sV', '-Pn', '-T0', '--scan-delay', '1s', '--max-retries', '5', '--data-length', '32']
    elif profile == 'decoy':
        # Decoy requires you to pick decoy IPs; here we include example RFC1918-like decoys — user should edit if they want real decoys.
        # NOTE: using decoy or source spoofing can be illegal/blocked and may require special network config.
        decoys = '192.0.2.1,198.51.100.2'  # example TEST-NET addresses; change as needed
        base = ['-sS', '-sV', '-Pn', '-T2', '--data-length', '24', '--decoy', decoys]
    else:
        base = ['-sV']

    if ports:
        base += ['-p', ports]
    base += [host]
    return base

def run_nmap_programmatic(host: str, ports: Optional[str], profile: str, timeout: int = 90) -> str:
    """
    Attempt to run nmap via python-nmap (nmap.PortScanner).
    Fallback to shell if python-nmap fails.
    Returns raw nmap output (stdout/stderr).
    """

    args = build_nmap_args(host, ports, profile)

    # python-nmap branch
    if nmap is not None:
        try:
            nm = nmap.PortScanner()

            # ---------------------------------------------------------
            # FIX: Remove any -p or port arguments from "arguments"
            # ---------------------------------------------------------
            clean_args = []
            skip_next = False

            for a in args:
                if skip_next:
                    skip_next = False
                    continue

                # Drop '-p' and the following port range
                if a == "-p":
                    skip_next = True
                    continue
                if a.startswith("-p"):
                    continue

                # Drop host (python-nmap provides host separately)
                if a == host:
                    continue

                clean_args.append(a)

            arguments = " ".join(clean_args)

            # Python-nmap requires only the PORT STRING here
            ports_arg = ports if ports else ""

            nm.scan(
                hosts=host,
                ports=ports_arg,
                arguments=arguments
            )

            # Output formatting
            out = (
                json.dumps(nm.get_nmap_last_output(), indent=2, default=str)
                if hasattr(nm, "get_nmap_last_output")
                else str(nm.scaninfo())
            )

            out += "\n\n-- parsed scan result (json-like) --\n" + json.dumps(
                nm._scan_result if hasattr(nm, "_scan_result") else {},
                indent=2,
                default=str,
            )

            return out

        except Exception as e:
            return (
                f"python-nmap error: {e}\n\n"
                "Falling back to shell nmap execution.\n\n"
                + run_nmap_shell(args, timeout=timeout)
            )

    # fallback
    return run_nmap_shell(args, timeout=timeout)


def run_nmap_shell(args: List[str], timeout: int = 90) -> str:
    nmap_bin = shutil.which("nmap")
    if not nmap_bin:
        return "nmap binary not found on PATH."
    cmd = [nmap_bin] + args
    try:
        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=timeout)
        out = proc.stdout
        if proc.stderr:
            out += "\n\nSTDERR:\n" + proc.stderr
        return out
    except Exception as e:
        return f"nmap shell execution error: {e}"

# ---------------------------
# CVE lookup (heuristic)
# ---------------------------
def extract_product_candidates(scan_results: Dict[str,Any]) -> List[str]:
    """
    Heuristically extract product keywords from headers and nmap output.
    Return a list of short tokens to query CVE search API.
    """
    candidates = set()
    # from headers
    headers = scan_results.get('headers', {}) or {}
    for key in ('Server', 'server', 'X-Powered-By', 'x-powered-by'):
        v = headers.get(key) or headers.get(key.lower())
        if v:
            # Pull the first token like "Apache/2.4.41 (Ubuntu)" -> "Apache", "2.4.41", "Ubuntu"
            # We'll add both product and version tokens
            parts = re.split(r'[;()\/\s]+', v)
            for p in parts:
                if len(p) >= 3 and not p.isdigit():
                    candidates.add(p.strip().strip('/'))
    # from nmap output (string)
    nmap_out = scan_results.get('nmap_output', '')
    if isinstance(nmap_out, str):
        # find common patterns product/version: e.g., "Apache httpd 2.4.41", "OpenSSH 7.4"
        matches = re.findall(r'([A-Za-z0-9\-\_\.]{3,})[ /\-]+([0-9]+\.[0-9]+(?:\.[0-9]+)?)', nmap_out)
        for m in matches:
            candidates.add(m[0])
            candidates.add(m[1])
        # also try to find words like "nginx", "apache", "tomcat" etc.
        simple = re.findall(r'(nginx|apache|tomcat|openssh|openvpn|iis|php|mysql|postgresql|mariadb|iis)', nmap_out, flags=re.IGNORECASE)
        for s in simple:
            candidates.add(s)
    # filter and return
    cleaned = [c for c in candidates if len(c) >= 3]
    return list(cleaned)[:12]  # limit to 12 keywords to avoid excessive API calls

def query_cve_for_keyword(keyword: str, timeout: int = 10) -> List[Dict[str,Any]]:
    """
    Query CIRCL cve search API for keyword. The API returns a JSON list of matches.
    If API is unavailable, returns empty list.
    """
    try:
        url = CIRCL_CVE_SEARCH.format(urllib.parse.quote(keyword))
        r = requests.get(url, timeout=timeout)
        if r.status_code == 200:
            data = r.json()
            # data has 'results' or is a list depending on endpoint; for /search it returns list of CVEs
            if isinstance(data, dict) and 'results' in data:
                return data['results']
            if isinstance(data, list):
                return data
        return []
    except Exception:
        return []

def perform_cve_lookups(scan_results: Dict[str,Any]) -> Dict[str,List[Dict[str,Any]]]:
    """
    For extracted product keywords, run CVE lookups and return mapping keyword -> list of CVEs (trimmed).
    We'll include only the top N vulnerabilities per keyword to keep the report concise.
    """
    cves = {}
    candidates = extract_product_candidates(scan_results)
    for k in candidates:
        found = query_cve_for_keyword(k)
        # trim results: keep id, summary, cvss (if present), published date
        trimmed = []
        for item in found[:10]:
            if isinstance(item, dict):
                entry = {
                    'id': item.get('id') or item.get('Name') or item.get('cve'),
                    'summary': item.get('summary') or item.get('summary'),
                    'cvss': item.get('cvss'),
                    'Published': item.get('Published') or item.get('PublishedDate') or item.get('LastModified')
                }
                trimmed.append(entry)
            else:
                # if item is string (some endpoints), just include as id
                trimmed.append({'id': str(item)})
        if trimmed:
            cves[k] = trimmed
    return cves

# ---------------------------
# Report generation
# ---------------------------
import datetime, json, html

def generate_html_report(target: str, scan_results: Dict[str,Any], outfile: str):
    #now = datetime.datetime.utcnow().isoformat() + "Z"
    now = datetime.datetime.now();
    escape = html.escape  # shortcut

    html_out = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Web Scan Report - {escape(target)}</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;margin:18px}} 
pre{{background:#f4f4f4;padding:8px;border-radius:6px;white-space:pre-wrap}} 
table{{border-collapse:collapse}} 
th,td{{border:1px solid #ccc;padding:6px;text-align:left}}
</style>
</head><body>
<h1>Web Scan Report</h1>
<p><strong>Target:</strong> {escape(target)}</p>
<p><strong>Generated:</strong> {now}</p>
<hr/>
"""

    # --- HTTP BASIC ---
    http_basic = scan_results.get('http_basic', {})
    headers = scan_results.get('headers', {})

    html_out += "<h2>HTTP Info</h2>\n"
    if 'error' in http_basic:
        html_out += f"<p style='color:red'>Error: {escape(str(http_basic.get('error')))}</p>"
    else:
        html_out += f"<p>Status: {escape(str(http_basic.get('status_code')))} | Final URL: {escape(str(http_basic.get('final_url')))}</p>\n"
        html_out += "<h3>Headers</h3>\n<pre>" + escape(json.dumps(headers, indent=2, default=str)) + "</pre>\n"

    # --- SECURITY HEADERS ---
    html_out += "<h2>Security Headers</h2>\n"
    sec = scan_results.get('security_headers', {})
    html_out += "<h3>Present</h3><pre>" + escape(json.dumps(sec.get('present', {}), indent=2, default=str)) + "</pre>\n"
    html_out += "<h3>Missing</h3><pre>" + escape(json.dumps(sec.get('missing', []), indent=2, default=str)) + "</pre>\n"

    # --- TLS / SSL ---
    html_out += "<h2>TLS / SSL</h2>\n<pre>" + escape(json.dumps(scan_results.get('tls', {}), indent=2, default=str)) + "</pre>\n"

    # --- HTTP METHODS ---
    html_out += "<h2>HTTP Methods</h2>\n<pre>" + escape(json.dumps(scan_results.get('http_methods', {}), indent=2, default=str)) + "</pre>\n"

    # --- COMMON PATHS ---
    html_out += "<h2>Common Paths</h2>\n<table><tr><th>Path</th><th>Status</th><th>Interesting</th><th>Snippet/Error</th></tr>\n"
    for r in scan_results.get('paths', {}).get('results', []):
        safe_path = escape(str(r.get('path')))
        safe_status = escape(str(r.get('status')))
        safe_interesting = escape(str(r.get('interesting')))
        raw_snippet = (r.get('snippet') or r.get('error') or '')[:800]
        safe_snippet = escape(str(raw_snippet))

        html_out += f"<tr><td>{safe_path}</td><td>{safe_status}</td><td>{safe_interesting}</td><td><pre>{safe_snippet}</pre></td></tr>\n"
    html_out += "</table>\n"

    # --- DIRECTORY LISTING ---
    html_out += "<h2>Directory Listing Checks</h2>\n<pre>" + escape(json.dumps(scan_results.get('dir_listing', {}), indent=2, default=str)) + "</pre>\n"

    # --- NMAP OUTPUT ---
    if scan_results.get('nmap_output'):
        html_out += "<h2>Nmap Output</h2>\n<pre>" + escape(scan_results.get('nmap_output')[:40000]) + "</pre>\n"

    # --- VULNERABILITIES ---
    html_out += "<h2>Vulnerabilities (heuristic)</h2>\n"
    html_out += "<pre>" + escape(json.dumps(scan_results.get('vulnerabilities', []), indent=2, default=str)) + "</pre>\n"

    # --- CVE LOOKUP ---
    if scan_results.get('cve_lookup'):
        html_out += "<h2>CVE Lookup Results (heuristic)</h2>\n"
        for k, vals in scan_results['cve_lookup'].items():
            html_out += f"<h3>Keyword: {escape(str(k))}</h3>\n<ul>\n"
            for item in vals:
                html_out += "<li><strong>{}</strong> - {} (cvss: {})</li>\n".format(
                    escape(str(item.get('id'))),
                    escape(str((item.get('summary') or '')[:200])),
                    escape(str(item.get('cvss')))
                )
            html_out += "</ul>\n"

    html_out += "<hr/><p>Scan finished.</p></body></html>"

    with open(outfile, 'w', encoding='utf-8') as f:
        f.write(html_out)

    print(f"HTML report saved to {outfile}")


def generate_xlsx_report(target: str, scan_results: Dict[str,Any], outfile: str):
    if Workbook is None:
        raise RuntimeError("openpyxl required. Install with: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Target", target])
    ws.append(["Generated", datetime.datetime.utcnow().isoformat() + "Z"])
    ws.append([])
    http_basic = scan_results.get('http_basic', {})
    ws.append(["HTTP Status", http_basic.get('status_code')])
    ws.append(["Final URL", http_basic.get('final_url')])
    ws.append([])
    ws.append(["Headers"])
    for k,v in (scan_results.get('headers') or {}).items():
        ws.append([k, str(v)])
    ws.append([])
    ws.append(["Security headers - present"])
    for k,v in (scan_results.get('security_headers',{}).get('present',{})).items():
        ws.append([k, str(v)])
    ws.append([])
    ws.append(["Missing security headers"])
    for m in (scan_results.get('security_headers',{}).get('missing',[])):
        ws.append([m])
    ws.append([])
    ws.append(["TLS"])
    tls = scan_results.get('tls',{})
    for k,v in tls.items():
        ws.append([k, str(v)])
    ws.append([])
    ws.append(["HTTP Methods", json.dumps(scan_results.get('http_methods',{}), default=str)])
    ws.append([])
    ws.append(["Paths checked"])
    ws.append(["Path","Status","Interesting","Snippet/Error"])
    for r in scan_results.get('paths',{}).get('results',[]):
        ws.append([r.get('path'), str(r.get('status')), str(r.get('interesting')), (r.get('snippet') or r.get('error') or '')[:1000]])
    ws.append([])
    ws.append(["Directory listings"])
    for d in scan_results.get('dir_listing',{}).get('results',[]):
        ws.append([d.get('dir'), d.get('url'), str(d.get('listing', d.get('status','')))])
    ws.append([])
    ws.append(["Vulnerabilities (heuristic)"])
    for v in scan_results.get('vulnerabilities',[]):
        ws.append([v])
    if scan_results.get('nmap_output'):
        ws2 = wb.create_sheet("nmap_output")
        lines = scan_results.get('nmap_output').splitlines()
        for i, line in enumerate(lines, start=1):
            ws2.cell(row=i, column=1, value=line)
    if scan_results.get('cve_lookup'):
        ws3 = wb.create_sheet("cve_lookup")
        row = 1
        for k, vals in scan_results['cve_lookup'].items():
            ws3.cell(row=row, column=1, value=f"Keyword: {k}")
            row += 1
            for item in vals:
                ws3.cell(row=row, column=1, value=str(item.get('id')))
                ws3.cell(row=row, column=2, value=(item.get('summary') or '')[:1000])
                ws3.cell(row=row, column=3, value=str(item.get('cvss')))
                row += 1
            row += 1
    wb.save(outfile)
    print(f"XLSX report saved to {outfile}")

def generate_json_report(target: str, scan_results: Dict[str,Any], outfile: str):
    payload = {'target': target, 'generated': datetime.datetime.utcnow().isoformat() + "Z", 'results': scan_results}
    with open(outfile, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, default=str)
    print(f"JSON report saved to {outfile}")

# ---------------------------
# Orchestration
# ---------------------------
def perform_scan(target_url: str, output_file: str, fast: bool=False, use_nmap: bool=True, nmap_profile: str='normal') -> Dict[str,Any]:
    target = norm_url(target_url)
    parsed = urllib.parse.urlparse(target)
    host, port = parse_host_port(parsed)

    scan_results: Dict[str,Any] = {}

    # HTTP basic
    http_basic = check_headers_and_server(target)
    scan_results['http_basic'] = http_basic
    scan_results['headers'] = http_basic.get('headers', {})

    # security headers
    scan_results['security_headers'] = check_security_headers(scan_results['headers'])

    # http methods
    scan_results['http_methods'] = check_http_methods(target)

    # common paths
    scan_results['paths'] = check_common_paths(target, COMMON_SENSITIVE_PATHS, fast=fast)

    # dir listing
    scan_results['dir_listing'] = check_directory_listing(target)

    # TLS
    if parsed.scheme == 'https':
        scan_results['tls'] = tls_info(host, port)
    else:
        scan_results['tls'] = {'note': 'not HTTPS'}

    # nmap
    if use_nmap:
        ports = str(port) if port else '80,443,8080'
        nmap_out = run_nmap_programmatic(host, ports, nmap_profile)
        scan_results['nmap_output'] = nmap_out
    else:
        scan_results['nmap_output'] = None

    # heuristics
    vulnerabilities = []
    missing = scan_results['security_headers'].get('missing', [])
    if parsed.scheme == 'https' and 'Strict-Transport-Security' in missing:
        vulnerabilities.append("Missing Strict-Transport-Security (HSTS) header on HTTPS site.")
    if 'X-Content-Type-Options' in missing:
        vulnerabilities.append("Missing X-Content-Type-Options header (could allow MIME sniffing).")
    for r in scan_results['paths']['results']:
        p = r.get('path')
        if p and ('.env' in p or '.git' in p) and r.get('status') == 200:
            vulnerabilities.append(f"Sensitive file accessible at {r.get('url')}")
        if r.get('status') == 200 and ('Index of /' in (r.get('snippet') or '')):
            vulnerabilities.append(f"Directory listing at {r.get('url')}")

    # TLS cipher check (basic)
    tls = scan_results.get('tls', {})
    cipher = tls.get('cipher')
    if cipher and isinstance(cipher, tuple):
        name = cipher[0]
        if any(k in name.lower() for k in ('rc4', 'des', '3des', 'md5')):
            vulnerabilities.append(f"Weak TLS cipher: {cipher}")

    scan_results['vulnerabilities'] = vulnerabilities

    # CVE lookups
    try:
        cve_results = perform_cve_lookups(scan_results)
        scan_results['cve_lookup'] = cve_results
    except Exception as e:
        scan_results['cve_lookup_error'] = str(e)

    # summary
    summary = []
    if http_basic.get('server_header'):
        summary.append(f"Server header: {http_basic.get('server_header')}")
    if scan_results.get('tls') and isinstance(scan_results['tls'], dict) and scan_results['tls'].get('protocol'):
        summary.append(f"TLS: {scan_results['tls'].get('protocol')}")
    if scan_results.get('vulnerabilities'):
        summary.append(f"{len(scan_results['vulnerabilities'])} heuristic issues")
    scan_results['summary'] = '\n'.join(summary)

    # output
    ext = os.path.splitext(output_file)[1].lower()
    if ext in ('.html', '.htm'):
        generate_html_report(target, scan_results, output_file)
    elif ext in ('.xlsx', '.xls'):
        generate_xlsx_report(target, scan_results, output_file)
    elif ext == '.json':
        generate_json_report(target, scan_results, output_file)
    else:
        # fallback to JSON
        generate_json_report(target, scan_results, output_file + '.json')

    return scan_results

# ---------------------------
# CLI
# ---------------------------
def main():
    parser = argparse.ArgumentParser(prog='web_scan_enhanced', description='Enhanced lightweight web server scanner.')
    parser.add_argument('target', help='Target URL (http:// or https:// or hostname)')
    parser.add_argument('--output', '-o', required=True, help='Output file (report.html / report.xlsx / report.json)')
    parser.add_argument('--fast', action='store_true', help='Fast (fewer path checks) mode')
    parser.add_argument('--no-nmap', action='store_true', help='Disable nmap scan')
    parser.add_argument('--nmap-profile', choices=['normal','stealth','paranoid','decoy'], default='normal',
                        help='nmap profile to use (stealth/paranoid/decoy/normal)')
    args = parser.parse_args()

    print("WARNING: Only scan targets you own or are authorized to test. Proceeding...")

    try:
        results = perform_scan(args.target, args.output, fast=args.fast, use_nmap=(not args.no_nmap), nmap_profile=args.nmap_profile)
        print("Scan complete.")
        print("Summary:")
        print(results.get('summary', 'No summary available.'))
        if results.get('vulnerabilities'):
            print("\nPotential issues found (heuristic):")
            for v in results['vulnerabilities']:
                print(" - " + v)
    except Exception as e:
        print("Error running scan:", e)
        sys.exit(2)

if __name__ == '__main__':
    main()

